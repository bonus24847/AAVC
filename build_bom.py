"""Generate AAVC_BOM.xlsx — hardware bill of materials for the AAVC 2026
hexacopter build (EFT X6100)."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TH = "Tahoma"  # renders Thai reliably

wb = Workbook()
ws = wb.active
ws.title = "BOM"

# ---- styles ----------------------------------------------------------------
hdr_fill = PatternFill("solid", fgColor="1F4E78")
cat_fill = PatternFill("solid", fgColor="DDEBF7")
own_fill = PatternFill("solid", fgColor="E2EFDA")
opt_fill = PatternFill("solid", fgColor="FFF2CC")
tot_fill = PatternFill("solid", fgColor="FCE4D6")
hdr_font = Font(name=TH, bold=True, color="FFFFFF", size=11)
cat_font = Font(name=TH, bold=True, size=11)
base_font = Font(name=TH, size=10)
bold_font = Font(name=TH, bold=True, size=10)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
baht = u'฿#,##0'

cols = ["หมวด", "รายการ", "รุ่นแนะนำ", "ทางเลือกประหยัด", "จำนวน",
        "ราคา/หน่วย (ประหยัด)", "ราคา/หน่วย (แนะนำ)", "รวม (ประหยัด)",
        "รวม (แนะนำ)", "สถานะ", "หมายเหตุ"]
widths = [14, 28, 26, 24, 7, 13, 13, 12, 12, 10, 40]

# ---- title -----------------------------------------------------------------
ws.merge_cells("A1:K1")
c = ws["A1"]
c.value = "AAVC 2026 — Hardware BOM (EFT X6100 hexacopter / 6S / ใบ 16\") · rules V1.3"
c.font = Font(name=TH, bold=True, size=14, color="FFFFFF")
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 24

ws.merge_cells("A2:K2")
c = ws["A2"]
c.value = ("AUW เต็ม payload ≈ 7.17 kg (Power-System-Guide-1.pdf — ชั่งจริงที่ G5)  ·  "
           "payload = ไข่ no.0 ในถุงหัวใจ ~0.1-0.15 kg "
           "(กติกา V1.3: 1 ฟอง/เที่ยว)  ·  ราคาเป็นประมาณการตลาด  ·  THB")
c.font = Font(name=TH, italic=True, size=9, color="808080")
c.alignment = Alignment(horizontal="center")

# ---- header row ------------------------------------------------------------
hr = 3
for i, name in enumerate(cols, start=1):
    cell = ws.cell(row=hr, column=i, value=name)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
ws.row_dimensions[hr].height = 30

row = hr + 1

def cat_header(title):
    global row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = cat_font
    cell.fill = cat_fill
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = border
    row += 1

def item(cat, name, rec, alt, qty, ulow, uhigh, status, note, fill=None):
    global row
    vals = [cat, name, rec, alt, qty, ulow, uhigh, None, None, status, note]
    for col, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = base_font
        cell.border = border
        cell.alignment = Alignment(vertical="center",
                                   wrap_text=(col in (2, 3, 4, 11)),
                                   horizontal=("center" if col in (5, 10) else "left"))
        if fill:
            cell.fill = fill
    # line totals via formula
    if qty and ulow is not None:
        ws.cell(row=row, column=8, value=f"=E{row}*F{row}").number_format = baht
        ws.cell(row=row, column=9, value=f"=E{row}*G{row}").number_format = baht
    for col in (6, 7, 8, 9):
        ws.cell(row=row, column=col).number_format = baht
    row += 1
    return row - 1

# ---- OWNED -----------------------------------------------------------------
cat_header("✅ มีอยู่แล้ว — ไม่ต้องซื้อ (ราคาประมาณการเพื่อคิดต้นทุน all-in)")
owned = [
    ("Flight controller", "Pixhawk 6X (standard set)", 1, 13000, 16000,
     "ชุดมาตรฐานมี PM02 มาด้วย; ลำนี้ใช้ PM02D จ่ายเฉพาะ FC (PM03D เดิมพัง 2026-08-16)"),
    ("Companion", "Raspberry Pi CM4 (เช็ก eMMC/Lite)", 1, 2500, 4500, "ราคาแปรตาม RAM/eMMC"),
    ("โครง", "EFT X6100 hexacopter (wheelbase ~1160mm)", 1, 2500, 5000, "เฟรม hexa-X 6 แขน"),
    ("ใบพัด", "16\" (ชุดใช้งาน)", 3, 400, 700, "หน่วย=คู่ (CW+CCW) — hexa ใช้ 3 คู่"),
    ("แบตเตอรี่", "DXF 6S 7500mAh 140C LiPo", 2, 2800, 4500,
     "ต่อขนาน 2 ก้อน = 15,000mAh @6S (2026-07-25); +1.05kg → AUW 8.22kg"),
    ("RC + Telemetry", "RadioMaster TX16S mk2 + Nomad + DBR4 Gemini", 1, 4000, 6000,
     "รีโมท + TX dual-band ELRS + RX diversity (CRSF เข้า TELEM1)"),
]
owned_first = row
for cat, name, qty, ulow, uhigh, note in owned:
    item(cat, name, "", "", qty, ulow, uhigh, "มีแล้ว", note, fill=own_fill)
owned_last = row - 1

# ---- A. PROPULSION ---------------------------------------------------------
cat_header("A. Propulsion")
buy_first = row
item("Propulsion", "มอเตอร์ (16\" 6S)", "T-Motor MN4014 330KV",
     "SunnySky X4112S 340KV", 6, 1200, 2200, "ซื้อ",
     "6 ตัวสำหรับ hexa; 3.84kg/ตัว @18\" → T/W 3.2:1 ที่ AUW 7.17kg; เช็ก bolt pattern arm X6100")
item("Propulsion", "ESC 60A DShot", "T-Motor Alpha 60A",
     "Hobbywing XRotor 60A", 6, 900, 1500, "ซื้อ",
     "แยก 6 ตัว (ไม่ใช่ 4-in-1) — ระบายร้อน/ซ่อมง่ายกว่า และ hexa บินต่อได้ถ้าเสีย 1 ตัว")
item("Propulsion", "ใบ 16\" คาร์บอน (สำรอง)", "1655 carbon", "", 3, 400, 700, "ซื้อ",
     "หน่วย=คู่; เน้น efficiency")

# ---- B. POWER --------------------------------------------------------------
cat_header("B. Power")
item("Power", "Power module 6S", "Holybro PM02D", "PM02 V3", 1, 1200, 2000, "ซื้อ",
     "PM02D (ติดตั้ง 2026-08-20 แทน PM03D ที่พัง) จ่ายไฟ FC/avionics เท่านั้น — "
     "มอเตอร์กินจากบอร์ดแยกที่ FC มองไม่เห็น ⇒ BAT1_CAPACITY ต้องเป็น -1 (เกจแรงดันล้วน) "
     "ห้ามตั้ง capacity เด็ดขาด: กระแสที่ FC เห็น (~0.7 A) คือ avionics ไม่ใช่ ~35-43 A "
     "ของการบิน เกจนับ mAh จะโกหกแบบไม่มีวันหมด")
item("Power", "BEC 5V/5A (เลี้ยง CM4 แยก)", "Mateksys / Pololu", "", 1, 300, 500, "ซื้อ",
     "อย่าดึงไฟ CM4 จาก FC rail")
item("Power", "PDB / สาย / XT90 / connector", "—", "", 1, 500, 1000, "ซื้อ", "")
item("Power", "สาย Y ขนานแบต 2 ก้อน (XT90 บัดกรี)", "สายซิลิโคน 10AWG + XT90",
     "", 1, 200, 500, "ซื้อ",
     "2026-07-25 ต่อขนาน: ต้องทนกระแสรวมเต็ม (hover 35.6A, peak สูงกว่ามาก) และ "
     "บัดกรี/ล็อกให้แน่น — ถ้าก้อนหนึ่งหลุดกลางอากาศแล้วต่อกลับ inrush อาจ brown-out FC")
item("Power", "บอร์ดชาร์จขนาน 6S + เช็กเกอร์เซลล์", "parallel charge board 6S "
     "+ cell checker/IR meter", "", 1, 400, 900, "ซื้อ",
     "แบตขนานต้อง MATCH ก่อนเสียบทุกครั้ง (ต่างกัน ≤~0.05-0.1V/cell) และเป็นคู่ตายตัว "
     "อายุ/รอบชาร์จเท่ากัน — ไม่มี fuse/diode คั่น ถ้าเซลล์ก้อนหนึ่งช็อต อีกก้อนจะอัดเข้าไป")

# ---- C. NAVIGATION ---------------------------------------------------------
cat_header("C. Navigation")
item("Nav", "GPS + compass (no-RTK)", "Holybro M10 + mast", "", 1, 1500, 2000, "ซื้อ",
     "no-RTK พอ; กล้องคุมเมตรสุดท้าย")
item("Nav", "Rangefinder ลงล่าง (AGL)", "Benewake TFmini-S (12m, UART/I2C)",
     "", 1, 800, 1400, "ซื้อ",
     "ยืนยัน AGL ช่วง descend + touchdown gate; EKF2_RNG_CTRL=1 (ตั้ง SENS_TFMINI_CFG "
     "ตามพอร์ตที่เลือกตอน G5). ไม่ใช้ optical flow แล้ว (ตัดออก 2026-07-22)")

# ---- D. VISION -------------------------------------------------------------
cat_header("D. Vision (mission-critical)")
item("Vision", "CM4 carrier (ถ้ายังไม่มี)", "Waveshare CM4 carrier mini",
     "", 1, 800, 2000, "ซื้อ", "ต้องมี USB+UART+power")
item("Vision", "CM4 heatsink/fan", "—", "", 1, 150, 400, "ซื้อ",
     "classical CV รันต่อเนื่อง ต้องระบายร้อน")
item("Vision", "microSD 32GB (ถ้า CM4 Lite)", "—", "", 1, 250, 400, "ซื้อ", "")
item("Vision", "กล้อง NADIR (ตัวเดียว)", "Meige OV9281 USB UVC mono GS 1280x720",
     "", 1, 0, 0, "มีแล้ว",
     "ตัดสินใจ 2026-07-15: กล้องเดียว mono global-shutter; กว้าง 1280px "
     "→ decode ArUco 400mm ที่ sweep 12m; HFOV เลนส์วัดจริงแล้ว 2026-08-17 "
     "= 74.2° (marker 50mm ที่ 0.495m → 85.5px → fx 847)")
item("Vision", "Gimbal servo + mount (stabilized nadir)", "servo metal-gear แกน pitch "
     "+ เคสกล้องกันสั่น", "", 1, 300, 900, "ซื้อ",
     "PX4 mount driver (MNT_*) กดกล้องดิ่งตลอด; VERIFY-AT-G5: ทิศ/ช่วง servo, "
     "PWM, ไม่ชน servo ปล่อยไข่ AUX 1-4 (4 ช่อง) (กล้อง oblique เดิม = ตัดออก)")

# ---- E. EGG RELEASE (briefing 2026-07-24: 4 ฟอง/เที่ยว, ห้าม winch) ---------
cat_header("E. Egg release (briefing 2026-07-24 — ไข่ 4 ฟอง/เที่ยว, "
           "ปล่อยหลังแตะพื้นเท่านั้น)")
item("Drop", "Servo metal-gear (ประตูช่องไข่)", "MG90S", "", 4, 120, 200, "ซื้อ",
     "4 ตัว = 4 กลไกปล่อย INDEPENDENT (กติกา §7 ห้ามปล่อยพร้อมกัน) — "
     "payload_id 0..3 → AUX 4/1/2/3 บน 6X (as-wired 2026-08-15: หน้าซ้าย/"
     "ท้ายขวา/หน้าขวา/ท้ายซ้าย — drop_servo_channels=[4,1,2,3], "
     "drop_payload_count=4); +สำรอง 2 ตัวก็ดี. ⚠ rail AUX ของ 6X ไม่จ่ายไฟเอง "
     "ต้องป้อน 5V เข้า servo rail (MG90S 4 ตัว peak ~1A/ตัว)")
item("Drop", "Egg cradle / enclosed bay x4 (กันสั่น)", "พิมพ์ 3D + โฟม/สปริงนุ่ม",
     "", 1, 800, 1600, "ซื้อ",
     "หน่วย=ชุด 4 ช่อง (ไข่ 4 ฟองบินไปพร้อมกันเที่ยวเดียว). กติกา: ห้าม winch, "
     "ปล่อยเบา ๆ บนแพด (ไข่ต้องรอด); ประตูเปิดล่าง ไข่ไหลลงต่ำ ~5cm. "
     "⚠ วางไม่ให้บังกล้อง nadir + TFmini-S ที่มองลงแกนกลาง และแบ่ง CG กับแบต 2 ก้อน")

# ---- F. MOUNTING / SAFETY / MISC ------------------------------------------
cat_header("F. Mounting / Safety / Misc")
item("Misc", "แผ่น vibration-damping (กล้อง+FC)", "—", "", 1, 300, 600, "ซื้อ",
     "ภาพสั่น = mask แดงเพี้ยน + projection noise")
item("Misc", "Antenna mount แยก 900/2.4", "—", "", 2, 75, 150, "ซื้อ",
     "กัน 900MHz desense จาก CM4/ESC")
item("Misc", "Standoff/น็อต/zip tie/heatshrink", "—", "", 1, 400, 600, "ซื้อ", "")
buy_last = row - 1

# ---- OPTIONAL --------------------------------------------------------------
cat_header("🔧 Optional upgrade")
opt_row = item("Optional", "แบต 6S 7500mAh Li-ion", "molicel / Li-ion 6S",
               "", 1, 3000, 4500, "optional",
               "Wh/น้ำหนักดีกว่า LiPo ~20-30% → คลาย energy budget", fill=opt_fill)

# ---- TOTALS ----------------------------------------------------------------
def total_line(label, lo_formula, hi_formula):
    global row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=8, value=lo_formula).font = bold_font
    ws.cell(row=row, column=9, value=hi_formula).font = bold_font
    ws.cell(row=row, column=8).number_format = baht
    ws.cell(row=row, column=9).number_format = baht
    for col in (1, 8, 9):
        ws.cell(row=row, column=col).fill = tot_fill
        ws.cell(row=row, column=col).border = border
    return row

owned_row = total_line("มูลค่าของที่มีอยู่แล้ว",
                       f"=SUM(H{owned_first}:H{owned_last})",
                       f"=SUM(I{owned_first}:I{owned_last})")
total_row = total_line("รวม ที่ต้องซื้อ (ไม่รวม optional)",
                       f"=SUM(H{buy_first}:H{buy_last})",
                       f"=SUM(I{buy_first}:I{buy_last})")
allin_row = total_line("รวมทั้งโปรเจกต์ (มีแล้ว + ต้องซื้อ)",
                       f"=H{owned_row}+H{total_row}",
                       f"=I{owned_row}+I{total_row}")
total_line("รวมทั้งโปรเจกต์ + Optional (Li-ion)",
           f"=H{allin_row}+H{opt_row}",
           f"=I{allin_row}+I{opt_row}")

# ---- finishing -------------------------------------------------------------
ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False
ws.print_options.horizontalCentered = True
ws.page_setup.fitToWidth = 1

out = "AAVC_BOM.xlsx"
wb.save(out)
print("wrote", out, "| buy rows", buy_first, "-", buy_last, "| total row", total_row)
